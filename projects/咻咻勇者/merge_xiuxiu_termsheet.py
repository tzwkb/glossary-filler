"""Merge filled EN back into the workbook's 术语总表 sheet.

Reads 咻咻勇者_术语总表_filled.xlsx, writes EN into the workbook by `_row`
(original Excel row number). ZH is re-verified per cell before writing;
mismatches are skipped and reported. Workbook is backed up first.
Alt_EN_* columns stay in the filled file for review, not merged.
"""
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

FILLED = Path(__file__).parent / "output" / "咻咻勇者_术语总表_filled.xlsx"
WORKBOOK = Path(__file__).parent / "source" / "咻咻勇者--内部工作表格.xlsx"
SHEET = "术语总表"


def main():
    df = pd.read_excel(FILLED)
    df = df[df["EN"].notna() & (df["EN"].astype(str).str.strip() != "")]
    print(f"filled rows to merge: {len(df)}")

    backup = WORKBOOK.with_name(
        f"{WORKBOOK.stem}.bak-{datetime.now():%Y%m%d-%H%M}{WORKBOOK.suffix}"
    )
    shutil.copy2(WORKBOOK, backup)
    print(f"backup: {backup.name}")

    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET]
    header = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    zh_col, en_col = header["ZH"], header["EN"]

    written, mismatched, occupied = 0, 0, 0
    for _, r in df.iterrows():
        row = int(r["_row"])
        cell_zh = str(ws.cell(row, zh_col).value or "").strip()
        if cell_zh != str(r["ZH"]).strip():
            print(f"  [skip] row {row}: sheet ZH '{cell_zh}' != filled ZH '{r['ZH']}'")
            mismatched += 1
            continue
        cur = ws.cell(row, en_col).value
        if cur is not None and str(cur).strip() != "":
            occupied += 1
            continue
        ws.cell(row, en_col).value = str(r["EN"]).strip()
        written += 1

    wb.save(WORKBOOK)
    print(f"written: {written}, zh-mismatch skipped: {mismatched}, already-filled skipped: {occupied}")


if __name__ == "__main__":
    main()
